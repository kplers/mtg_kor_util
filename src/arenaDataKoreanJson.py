import sqlite3
import glob
import os
import sys
import json
import re
import unicodedata
import pandas as pd  # New import for Excel export
# New import for Excel export
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import subprocess
# Define the client folder path
CLIENT_FOLDER = r"C:\Program Files (x86)\Steam\steamapps\common\MTGA\MTGA_Data\Downloads\Raw"

# Define color order for sorting
COLOR_ORDER = {"W": 1, "U": 2, "B": 3, "R": 4, "G": 5}
COLOR_INDEX = {"1": "W", "2": "U", "3": "B", "4": "R", "5": "G"}

# Define rarity index for mapping rarity numbers to strings
RARITY_INDEX = {2: "커먼", 3: "언커먼", 4: "레어", 5: "미식레어"}
# 희귀도에 따른 색상 매핑 (Hex 코드)
RARITY_COLOR_MAPPING = {
    "커먼": "808080",
    "언커먼": "C0C0C0",
    "레어": "FFD700",
    "미식레어": "E06C00"
}


def color_sort_key(card):
    """
    Sorting key based on collector number
    """
    return int(card.get("collector_number", -1))


def convertMana(text):
    """
    Converts mana symbols from "oWoWoW" to "{W}{W}{W}" format within a given text.

    """
    if not text:
        return ""

    # Regular expression to find all "o.." patterns
    pattern = re.compile(r'o([^o]+)')

    # Function to process each match
    def replace_match(match):
        mana_cost = match.group(1)
        # Split by 'o' without stripping 'o's
        parts = mana_cost.split('o')
        # Filter out empty strings and wrap each mana symbol with braces
        converted = ''.join([f"{{{part.strip('(').strip(')')}}}" for part in parts if part])
        return converted

    # Substitute all matches in the text
    converted_text = pattern.sub(replace_match, text)
    return converted_text


def convertText(text):
    """
    Converts mana cost from "{oWoWoW}" to "{W}{W}{W}" format within a given text.
    Handles multiple {..} patterns in the text without removing braces.
    Also handles some edge cases like classes: 
        CLASSLEVEL \[.*?\] \[.*?\]\n => (blank)
        #ClassLevelArtifact\n => (blank)
    """
    if not text:
        return ""

    # Regular expression to find all {..} patterns
    pattern = re.compile(r'\{([^}]+)\}')

    # Function to process each match
    def replace_match(match):
        mana_cost = match.group(1)
        # Split by 'o' without stripping 'o's
        parts = mana_cost.split('o')
        # Filter out empty strings and wrap each mana symbol with braces
        converted = ''.join([f"{{{part}}}" for part in parts if part])
        return converted

    # Substitute all matches in the text
    converted_text = pattern.sub(replace_match, text)
    
    # Now do the edge cases
    converted_text = re.sub(r'CLASSLEVEL \[.*?\] \[.*?\]', '', converted_text)
    converted_text = re.sub(r'#ClassLevelArtifact', '', converted_text)
    return converted_text


def getManaValue(manaCost):
    """
    Calculates the mana value from the mana cost string.
    """
    return sum(int(c) if c.isdigit() else 1 for c in manaCost)


def normalize_text(text):
    return unicodedata.normalize('NFC', text)


def getSet(setCode, digital=False):
    """
    Processes the MTGA card database and exports the specified set to JSON and Excel files.
    """
    cards_ids = []  # List to store card IDs and details
    cards_json = []  # List to store the final card data

    db_path_pattern = os.path.join(CLIENT_FOLDER, "Raw_CardDatabase_*.mtga")
    db_files = glob.glob(db_path_pattern)

    if not db_files:
        print("No database files found.")
        return

    # Initialize localization dictionary
    localization_dict = {}
    # Initialize ability ID - text ID map
    ability_text_map = {}

    # First pass: Fetch all relevant localizations and abilities from all databases
    for db in db_files:
        try:
            with sqlite3.connect(db) as conn:
                cursor = conn.cursor()

                # Define localization table and columns
                localization_table = 'Localizations'
                loc_id_col = 'LocId'
                koKR_col = 'koKR'
                enUS_col = 'enUS'
                formatted_col = 'Formatted'

                cards_table = 'Cards'
                expansion_code_col = 'ExpansionCode' if not digital else 'DigitalReleaseSet'
                # Check if the current database contains the desired set
                check_query = f"""
                    SELECT COUNT(*)
                    FROM {cards_table}
                    WHERE {expansion_code_col} LIKE ?
                """
                cursor.execute(check_query, (setCode,))
                count = cursor.fetchone()[0]
                if count < 2:
                    continue  # Likely not the correct set, skip to next database
                # Fetch all localizations
                localization_query = f"""
                    SELECT {loc_id_col}, {koKR_col}, {enUS_col}, {formatted_col}
                    FROM {localization_table}
                """
                cursor.execute(localization_query)
                localizations = cursor.fetchall()

                for loc_id, koKR, enUS, formatted in localizations:
                    if formatted == "1":
                        continue  # Skip formatted entries
                    localization_dict[str(loc_id)] = {
                        "korean": normalize_text(koKR) if koKR else "(미완성)",
                        "english": normalize_text(enUS) if enUS else "(미완성)"
                    }

                # Fetch abilities
                id_col = "Id"
                text_col = "TextId"
                loyaltyCost_col = "LoyaltyCost"
                baseid_col = "BaseId"
                baseidnum_col = "BaseIdNumeral"
                ability_table = "Abilities"
                ability_query = f"""
                    SELECT {id_col}, {text_col}, {loyaltyCost_col}, {baseid_col}, {baseidnum_col}
                    FROM {ability_table}
                """
                cursor.execute(ability_query)
                abilities = cursor.fetchall()
                for ability_id, text_id, loyaltyCost, baseid, baseidnum in abilities:
                    ability_text_map[str(ability_id)] = (str(text_id), str(loyaltyCost), str(baseid), str(baseidnum))

        except sqlite3.Error as e:
            print(f"Error fetching localizations from {db}: {e}")
            continue  # Skip to the next database if error occurs

    # Second pass: Process cards from relevant databases
    for db in db_files:
        try:
            with sqlite3.connect(db) as conn:
                cursor = conn.cursor()

                # Define table and column names
                cards_table = 'Cards'
                expansion_code_col = 'ExpansionCode' if not digital else 'DigitalReleaseSet'

                # Columns to select from Cards table
                card_columns = [
                    "TitleId", "AltTitleId", "AbilityIds", "FlavorTextId",
                    "ReminderTextId", "TypeTextId", "SubtypeTextId", "Rarity",
                    "CollectorNumber", "OldSchoolManaText", "Power",
                    "Toughness", "Types", "Colors", "IsPrimaryCard"
                ]
                columns_str = ", ".join(card_columns)

                # Check if the current database contains the desired set
                check_query = f"""
                    SELECT COUNT(*)
                    FROM {cards_table}
                    WHERE {expansion_code_col} LIKE ?
                """
                cursor.execute(check_query, (setCode,))
                count = cursor.fetchone()[0]
                if count < 2:
                    continue  # Likely not the correct set, skip to next database

                # Fetch all card data for the set in one query
                fetch_query = f"""
                    SELECT {columns_str}
                    FROM {cards_table}
                    WHERE {expansion_code_col} LIKE ?
                """
                cursor.execute(fetch_query, (setCode,))
                rows = cursor.fetchall()

                # Define index mapping
                INDEX = [
                    "title_id", "alt_title_id", "ability_id", "flavor_text_id",
                    "reminder_text_id", "type_id", "subtype_id",
                    "rarity", "collector_number", "mana_text", "power",
                    "toughness", "type", "color", "is_primary_card"
                ]

                # Process each card
                for row in rows:
                    card = {}
                    card_valid=True
                    for i, value in enumerate(row):
                        key = INDEX[i]
                        if key=="is_primary_card":
                            if value==0 and digital:
                                card_valid=False
                                break
                            continue
                        if value not in (None, "", "0"):
                            card[key] = value
                        elif key in ["power", "toughness"] and value == "0":
                            card[key] = value
                    if card_valid:
                        cards_ids.append(card)
        except sqlite3.Error as e:
            print(f"Error processing cards from {db}: {e}")
            continue  # Skip to the next database if error occurs

    # Now process each card for localization and other fields
    for card in cards_ids:
        card_json = {}
        validCard = True
        if "rarity" not in card:
            continue
        for key, value in card.items():
            value = str(value)
            if value in [None, "None", "0", ""]:
                if key == "rarity" or key == "type_id":
                    validCard = False
                    break
                continue

            if key in ["title_id", "alt_title_id", "flavor_text_id", "reminder_text_id", "type_id", "subtype_id"]:
                if key == "flavor_text_id" and value in ["0", "1"]:
                    continue
                loc_data = localization_dict.get(
                    str(value), {"korean": "(미완성)", "english": "(미완성)"})
                if key == "type_id" and "대지" in loc_data["korean"]:
                    loc_data["korean"] = "대지"
                field_name = key.replace('_id', '')
                card_json[f"korean_{field_name}"] = loc_data["korean"]
                card_json[f"english_{field_name}"] = loc_data["english"]
            elif key == "ability_id":
                korean_ability = []
                english_ability = []
                abilities = value.split(",")
                for ability in abilities:
                    ability = ability.strip()
                    if ':' in ability:
                        # Handle abilities with ':' separator
                        left=ability.split(":")[0]
                        ability = ability.split(":")[-1]
                        left_ability=list(ability_text_map.get(left, ("0", "", "", "")))
                        loc_data = localization_dict.get(
                            ability, {"korean": "(미완성)", "english": "(미완성)"})
                        left_ability[0]=ability
                        ability=tuple(left_ability)
                    else:
                        # Handle abilities without ':' separator
                        ability = (ability_text_map.get(ability, ("0", "", "", "")))
                        
                        loc_data = localization_dict.get(
                            ability[0], {"korean": "(미완성)", "english": "(미완성)"})
                    # Apply convertText to ability texts
                    formatted_korean = convertText(loc_data["korean"])
                    formatted_english = convertText(loc_data["english"])
                    
                    # Planeswalkers
                    if ability[1]!="" and ability[1]!=None and ability[1]!="None":
                        formatted_korean = ability[1]+": "+formatted_korean
                        formatted_english = ability[1]+": "+formatted_english
                    # Sagas
                    elif ability[2]=="166" and ability[3]!="0":
                        saga_text={"1":"I", "2":"II", "3":"III", "4":"IV", "5":"V", "6":"VI", "7":"VII", "8":"VIII", "9":"IX", "10":"X"}
                        try:
                            kor_ability_split=korean_ability[-1].split(":")
                            eng_ability_split=english_ability[-1].split(":")
                            if formatted_korean.strip()==kor_ability_split[1].strip():
                                korean_ability[-1]=kor_ability_split[0].strip()+", "+saga_text[ability[3]]+": "+formatted_korean
                                english_ability[-1]=eng_ability_split[0].strip()+", "+saga_text[ability[3]]+": "+formatted_english
                                formatted_korean=''
                                formatted_english=''
                                continue
                        except:
                            pass
                        
                        finally:
                            formatted_korean=saga_text[ability[3]]+": "+formatted_korean
                            formatted_english=saga_text[ability[3]]+": "+formatted_english
                    if formatted_korean!='' and formatted_english!='':
                        korean_ability.append(formatted_korean)
                        english_ability.append(formatted_english)
        
                card_json["korean_ability"] = "\n".join(korean_ability)
                card_json["english_ability"] = "\n".join(english_ability)
            elif key == "toughness":
                if "power" not in card:
                    if card.get("type") == "8":
                        card_json["loyalty"] = value
                    else:
                        card_json["defense"] = value
                else:
                    card_json["toughness"] = card["toughness"]
            elif key == "color":
                colors = value.split(",")
                col = "".join([COLOR_INDEX.get(c.strip(), "") for c in colors])
                card_json["color"] = col
            elif key == "mana_text":
                card_json["mana_cost"] = convertMana(value)
            elif key == "rarity":
                try:
                    rarity = int(value)
                    if rarity not in RARITY_INDEX:
                        validCard = False
                        break
                    card_json["rarity"] = RARITY_INDEX[rarity]
                except ValueError:
                    validCard = False
                    break
            elif key == "type":
                continue  # Skip type field
            else:
                card_json[key] = value
        if validCard and (card_json["korean_title"] not in [card["korean_title"] for card in cards_json]):
            cards_json.append(card_json)

    # Prepare data for Excel export
    # Convert list of dictionaries to pandas DataFrame
    df = pd.DataFrame(cards_json)
    desired_column_order = [
        "collector_number",
        "color",
        "mana_cost",
        "korean_title",
        "english_title",
        "korean_type",
        "english_type",
        "korean_subtype",
        "english_subtype",
        "korean_ability",
        "english_ability",
        "korean_flavor_text",
        "english_flavor_text",
        "power",
        "toughness",
        "loyalty",
        "korean_reminder_text",
        "english_reminder_text",
        "rarity"
    ]
    cards_json.sort(key=color_sort_key)
    # Convert list of dictionaries to pandas DataFrame
    df = pd.DataFrame(cards_json)

    # Reorder columns based on desired order, ignoring missing columns
    existing_columns = [
        col for col in desired_column_order if col in df.columns]
    df = df[existing_columns]

    

    # Define output directory
    output_dir = os.path.join("..", "data")
    os.makedirs(output_dir, exist_ok=True)

    # Define file paths
    json_output_path = os.path.join(output_dir, f"{setCode}.json")
    excel_output_path = os.path.join(output_dir, f"{setCode}.xlsx")
    ods_output_path = os.path.join(output_dir, f"{setCode}.ods")  # New ODS file path
    # Export to JSON
    with open(json_output_path, 'w', encoding='utf-8') as f:
        json.dump(cards_json, f, indent=4, sort_keys=True, ensure_ascii=False)

    # Export to Excel with formatting
    with pd.ExcelWriter(excel_output_path, engine='openpyxl') as writer:
        # Define column widths (customize as needed)
        column_widths = {
            "#": 5,                   # collector_number
            "색": 5,                 # color
            "마나 비용": 10,           # mana_cost
            "한글명": 13,             # korean_title
            "영문명": 13,             # english_title
            "한글 유형": 12,           # korean_type
            "영문 유형": 12,           # english_type
            "한글 하위 유형": 15,       # korean_subtype
            "영문 하위 유형": 15,       # english_subtype
            "한글 능력": 25,           # korean_ability
            "영어 능력": 25,           # english_ability
            "한글 플레이버 텍스트": 22,  # korean_flavor_text
            "영어 플레이버 텍스트": 22,  # english_flavor_text
            "공": 5,                  # power
            "방": 5,                  # toughness
            "충": 5,                  # loyalty
            "수": 5,                  # defense
            "한글 리마인더 텍스트": 20,  # korean_reminder_text
            "영어 리마인더 텍스트": 20,  # english_reminder_text
            "희귀도": 7             # rarity
        }

        # 열 이름 매핑 사전 정의
        column_name_mapping = {
            "collector_number": "#",
            "color": "색",
            "mana_cost": "마나 비용",
            "korean_title": "한글명",
            "english_title": "영문명",
            "korean_type": "한글 유형",
            "english_type": "영문 유형",
            "korean_subtype": "한글 하위 유형",
            "english_subtype": "영문 하위 유형",
            "rarity": "희귀도",
            "korean_ability": "한글 능력",
            "english_ability": "영어 능력",
            "korean_flavor_text": "한글 플레이버 텍스트",
            "english_flavor_text": "영어 플레이버 텍스트",
            "power": "공",
            "toughness": "방",
            "loyalty": "충",
            "defense": "수",
            "korean_reminder_text": "한글 리마인더 텍스트",
            "english_reminder_text": "영어 리마인더 텍스트"
        }

        # Rename columns to Korean
        df.rename(columns=column_name_mapping, inplace=True)

        # Write DataFrame to Excel
        df.to_excel(writer, index=False, sheet_name='Cards')
        worksheet = writer.sheets['Cards']

        # 열 이름과 열 번호 매핑 생성
        header = next(worksheet.iter_rows(
            min_row=1, max_row=1, values_only=True))
        column_mapping = {header[i]: i+1 for i in range(len(header))}

        # Apply column widths based on column names
        for col_name, width in column_widths.items():
            col_num = column_mapping.get(col_name)
            if col_num:
                col_letter = get_column_letter(col_num)
                worksheet.column_dimensions[col_letter].width = width

        # Apply text wrapping and center alignment to all cells
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = Alignment(
                    wrap_text=True, horizontal='center', vertical='center')

        # Apply font to all cells
        korean_font = Font(name='맑은 고딕')
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = korean_font

        # Add hyperlinks to "한글명" and "영문명" columns
        set_code_lower=""
        if not digital:
            set_code_lower = setCode.lower()
        elif setCode.split("-")[0][0]=="Y":
            # Alchemy Set
            set_code_lower="y"+setCode.split("-")[1].lower()
        elif setCode.split("-")[0][0]=="S":
            # Special Guest Set
            set_code_lower="spg"
        for row in range(2, worksheet.max_row + 1):
            collector_number = worksheet.cell(
                row=row, column=column_mapping.get("#", 1)).value
            if not collector_number:
                continue  # Skip if collector_number is missing
            scryfall_url = f"https://scryfall.com/card/{set_code_lower}/{collector_number}"

            # 한글명 (D열) 하이퍼링크 추가
            korean_col_num = column_mapping.get("한글명")
            if korean_col_num:
                korean_name_cell = worksheet.cell(
                    row=row, column=korean_col_num)
                korean_name_cell.hyperlink = scryfall_url
                korean_name_cell.style = "Hyperlink"
                # Reapply alignment and bold font after setting hyperlink
                korean_name_cell.alignment = Alignment(
                    wrap_text=True, horizontal='center', vertical='center')
                korean_name_cell.font = Font(name='맑은 고딕', bold=True)

            # 영문명 (E열) 하이퍼링크 추가
            english_col_num = column_mapping.get("영문명")
            if english_col_num:
                english_name_cell = worksheet.cell(
                    row=row, column=english_col_num)
                english_name_cell.hyperlink = scryfall_url
                english_name_cell.style = "Hyperlink"
                # Reapply alignment and bold font after setting hyperlink
                english_name_cell.alignment = Alignment(
                    wrap_text=True, horizontal='center', vertical='center')
                english_name_cell.font = Font(name='맑은 고딕', bold=True)

        # Identify "희귀도" column dynamically
        rarity_col_num = column_mapping.get("희귀도")
        collector_number_col_num = column_mapping.get("#")
        if rarity_col_num:
            # Apply fill color based on rarity
            for row in range(2, worksheet.max_row + 1):
                rarity = worksheet.cell(row=row, column=rarity_col_num).value
                if not rarity:
                    continue  # Skip if rarity is missing
                fill_color = RARITY_COLOR_MAPPING.get(
                    rarity, "FFFFFF")  # Default to white if not found
                fill = PatternFill(start_color=fill_color,
                                   end_color=fill_color, fill_type="solid")
                worksheet.cell(row=row, column=rarity_col_num).fill = fill
                worksheet.cell(row=row, column=collector_number_col_num).fill = fill
    try:
        # Check if LibreOffice is installed by attempting to find 'soffice'
        from shutil import which
        soffice_path = which('soffice')
        if soffice_path is None:
            print("LibreOffice 'soffice' executable not found. Please install LibreOffice and ensure 'soffice' is in your PATH.")
        else:
            # Convert XLSX to ODS using LibreOffice
            convert_xlsx_to_ods(excel_output_path, ods_output_path, libreoffice_path=soffice_path)
    except Exception as e:
        print(f"An error occurred during ODS export: {e}")

    print(f"Successfully exported {setCode}.json, {setCode}.xlsx, and {setCode}.ods")
    return 0
def convert_xlsx_to_ods(xlsx_path, ods_path, libreoffice_path=r"C:\Program Files\LibreOffice\program\soffice.exe"):
    """
    Converts an XLSX file to ODS using LibreOffice's command-line tool.

    Parameters:
    - xlsx_path: Path to the input XLSX file.
    - ods_path: Path to the output ODS file.
    - libreoffice_path: Path to the LibreOffice executable ('soffice').
                        If 'soffice' is in PATH, no need to specify full path.
    """
    try:
        # Command to convert XLSX to ODS
        # --headless: Run without GUI
        # --convert-to: Specify the output format
        # --outdir: Specify the output directory
        # The actual output file name will be set by LibreOffice
        output_dir = os.path.dirname(ods_path)
        command = [
            libreoffice_path,
            '--headless',
            '--convert-to', 'ods',
            '--outdir', output_dir,
            xlsx_path
        ]
        subprocess.run(command, check=True)
        # Rename the converted file to the desired ODS file name
        base_name = os.path.splitext(os.path.basename(xlsx_path))[0]
        generated_ods = os.path.join(output_dir, f"{base_name}.ods")
        if os.path.exists(generated_ods):
            os.rename(generated_ods, ods_path)
            print(f"Converted {xlsx_path} to {ods_path} successfully.")
        else:
            print(f"Conversion failed: {generated_ods} not found.")
    except subprocess.CalledProcessError as e:
        print(f"LibreOffice conversion failed: {e}")
    except Exception as e:
        print(f"Error during XLSX to ODS conversion: {e}")

